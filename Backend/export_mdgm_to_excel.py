"""
Export MDGM seed data to Excel for review.
Run from Backend folder: python export_mdgm_to_excel.py
Creates MDGM_Seed_Data.xlsx in the Backend folder.
"""
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
except ImportError:
    print("Run: pip install openpyxl")
    sys.exit(1)

from seed_data_mdgm import REGIONS, COUNTRIES, BRANDS_TA, SKUS, MDGM_ROWS


def main():
    wb = Workbook()

    # Sheet 1: MDGM rows (~100)
    ws_mdgm = wb.active
    ws_mdgm.title = "MDGM"
    headers = [
        "sku_id", "country", "region", "therapeutic_area", "brand", "channel",
        "price_type", "current_price_eur", "marketed_status", "currency",
    ]
    for col, h in enumerate(headers, 1):
        ws_mdgm.cell(row=1, column=col, value=h)
        ws_mdgm.cell(row=1, column=col).font = Font(bold=True)
    for row_idx, row in enumerate(MDGM_ROWS, 2):
        for col_idx, val in enumerate(row, 1):
            ws_mdgm.cell(row=row_idx, column=col_idx, value=val)
    ws_mdgm.column_dimensions["A"].width = 18
    ws_mdgm.column_dimensions["C"].width = 10
    ws_mdgm.column_dimensions["D"].width = 18
    ws_mdgm.column_dimensions["E"].width = 14

    # Sheet 2: Regions
    ws_reg = wb.create_sheet("Regions", 1)
    ws_reg.cell(row=1, column=1, value="code").font = Font(bold=True)
    ws_reg.cell(row=1, column=2, value="name").font = Font(bold=True)
    for row_idx, (code, name) in enumerate(REGIONS, 2):
        ws_reg.cell(row=row_idx, column=1, value=code)
        ws_reg.cell(row=row_idx, column=2, value=name)

    # Sheet 3: Countries
    ws_cnt = wb.create_sheet("Countries", 2)
    ws_cnt.cell(row=1, column=1, value="code").font = Font(bold=True)
    ws_cnt.cell(row=1, column=2, value="name").font = Font(bold=True)
    ws_cnt.cell(row=1, column=3, value="region").font = Font(bold=True)
    for row_idx, (code, name, region) in enumerate(COUNTRIES, 2):
        ws_cnt.cell(row=row_idx, column=1, value=code)
        ws_cnt.cell(row=row_idx, column=2, value=name)
        ws_cnt.cell(row=row_idx, column=3, value=region)

    # Sheet 4: Brand -> Therapeutic area
    ws_ta = wb.create_sheet("Brand_TA", 3)
    ws_ta.cell(row=1, column=1, value="brand").font = Font(bold=True)
    ws_ta.cell(row=1, column=2, value="therapeutic_area").font = Font(bold=True)
    for row_idx, (brand, ta) in enumerate(BRANDS_TA, 2):
        ws_ta.cell(row=row_idx, column=1, value=brand)
        ws_ta.cell(row=row_idx, column=2, value=ta)

    # Sheet 5: SKUs
    ws_sku = wb.create_sheet("SKUs", 4)
    ws_sku.cell(row=1, column=1, value="sku_id").font = Font(bold=True)
    ws_sku.cell(row=1, column=2, value="name").font = Font(bold=True)
    for row_idx, (sku_id, name) in enumerate(SKUS, 2):
        ws_sku.cell(row=row_idx, column=1, value=sku_id)
        ws_sku.cell(row=row_idx, column=2, value=name)

    out_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "MDGM_Seed_Data.xlsx")
    wb.save(out_path)
    print(f"Saved: {out_path}")
    print(f"  Sheet 'MDGM': {len(MDGM_ROWS)} rows")
    print(f"  Sheet 'Regions': {len(REGIONS)} rows")
    print(f"  Sheet 'Countries': {len(COUNTRIES)} rows")
    print(f"  Sheet 'Brand_TA': {len(BRANDS_TA)} rows")
    print(f"  Sheet 'SKUs': {len(SKUS)} rows")


if __name__ == "__main__":
    main()
